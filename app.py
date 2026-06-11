import io
import re
import zipfile
import tempfile
from copy import copy
from datetime import datetime, date
from pathlib import Path
from html import escape

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
except Exception:
    colors = None

PAS_YELLOW = "#FFD400"
PAS_BLACK = "#0A0A0A"
PAS_DARK = "#171717"
PAS_GREY = "#F4F4F4"
APP_VERSION = "v1.0 Prototype Build"
TEMPLATE_PATHS = [
    Path("assets/Hire Report Template.xlsx"),
    Path("Hire Report Template.xlsx"),
]

st.set_page_config(page_title="PAS Vendor On-Hire Checker", page_icon="pas_logo.png", layout="wide")

st.markdown(
    f"""
    <style>
    .stApp {{ background: #f5f5f5; color: #0A0A0A; }}
    section[data-testid="stSidebar"] {{
        background: {PAS_BLACK};
        color: white;
        padding-top: 1.45rem;
    }}
    section[data-testid="stSidebar"] * {{ color: white; }}
    section[data-testid="stSidebar"] img {{
        margin-top: 0.15rem;
        border-radius: 14px;
    }}
    .block-container {{
        padding-top: 1.4rem;
        padding-bottom: 2rem;
        max-width: 1500px;
    }}

    .pas-hero {{
        background: linear-gradient(135deg, {PAS_BLACK} 0%, #202020 70%, #7a6900 135%);
        border-radius: 18px;
        padding: 24px 28px;
        margin-bottom: 18px;
        box-shadow: 0 8px 25px rgba(0,0,0,0.12);
    }}
    .pas-title {{
        color: white;
        font-size: 32px;
        font-weight: 900;
        margin: 0;
        letter-spacing: -0.03em;
    }}
    .pas-subtitle {{
        color: {PAS_YELLOW};
        font-size: 14px;
        margin-top: 4px;
        font-weight: 800;
    }}

    .kpi-card {{
        background: white;
        border-radius: 18px;
        padding: 18px 20px;
        border: 1px solid #e8e8e8;
        box-shadow: 0 3px 12px rgba(0,0,0,0.05);
        min-height: 112px;
    }}
    .kpi-label {{
        color: #111;
        font-size: 14px;
        font-weight: 800;
        margin-bottom: 8px;
    }}
    .kpi-value {{
        color: {PAS_YELLOW};
        font-size: 36px;
        font-weight: 950;
        line-height: 1.05;
        text-shadow: 0 1px 0 #111;
    }}
    .kpi-sub {{
        color: #222;
        font-size: 13px;
        margin-top: 6px;
    }}

    .stButton > button, .stDownloadButton > button {{
        background: {PAS_YELLOW} !important;
        color: {PAS_BLACK} !important;
        border: 1px solid {PAS_BLACK} !important;
        border-radius: 12px !important;
        font-weight: 900 !important;
    }}

    .stCaption, div[data-testid="stCaptionContainer"], .stMarkdown p, .stInfo {{
        color: #0A0A0A !important;
    }}

    .pas-results-title {{
        color: #0A0A0A;
        font-size: 26px;
        font-weight: 950;
        margin: 22px 0 8px 0;
    }}
    .pas-unmatched-pill {{
        display: inline-flex;
        align-items: center;
        gap: 8px;
        background: {PAS_YELLOW};
        color: {PAS_BLACK};
        border: 1px solid #111;
        border-radius: 14px 14px 0 0;
        padding: 11px 18px;
        font-weight: 950;
        box-shadow: 0 3px 10px rgba(0,0,0,0.08);
        margin-top: 4px;
    }}

    .pas-table-wrap {{
        background: white;
        border: 1px solid #d9d9d9;
        border-radius: 0 16px 16px 16px;
        overflow: auto;
        box-shadow: 0 4px 18px rgba(0,0,0,0.07);
        margin-bottom: 18px;
    }}
    table.pas-table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 13px;
        color: #0A0A0A;
        background: white;
    }}
    table.pas-table thead th {{
        background: {PAS_YELLOW};
        color: {PAS_BLACK};
        font-weight: 950;
        text-align: left;
        padding: 11px 12px;
        border: 1px solid #c7a900;
        white-space: nowrap;
    }}
    table.pas-table tbody td {{
        background: white;
        color: #0A0A0A;
        padding: 9px 12px;
        border: 1px solid #e3e3e3;
        vertical-align: top;
    }}
    table.pas-table tbody tr:nth-child(even) td {{ background: #fbfbfb; }}
    table.pas-table a {{ color: #006fd6 !important; font-weight: 800; text-decoration: none; }}
    .pas-note {{ color: #0A0A0A; font-size: 13px; margin: 8px 0 16px 0; }}
    .pas-support {{ color: #0A0A0A; font-size: 14px; margin: 16px 0; }}

    div[data-testid="stAlert"], div[data-testid="stAlert"] * {{ color: #0A0A0A !important; }}
    div[data-testid="stAlert"] {{ border: 1px solid #e2ba00 !important; }}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    f"""
    <style>
    section[data-testid="stSidebar"] .stMarkdown p,
    section[data-testid="stSidebar"] .stMarkdown li,
    section[data-testid="stSidebar"] .stMarkdown h1,
    section[data-testid="stSidebar"] .stMarkdown h2,
    section[data-testid="stSidebar"] .stMarkdown h3,
    section[data-testid="stSidebar"] .stMarkdown strong,
    section[data-testid="stSidebar"] .stMarkdown span {{ color: #ffffff !important; }}

    div[data-testid="stFileUploader"] svg,
    div[data-testid="stFileUploader"] button svg,
    div[data-testid="stFileUploader"] [data-testid="stIconMaterial"] {{
        color: #FFD400 !important;
        fill: #FFD400 !important;
        stroke: #FFD400 !important;
    }}
    div[data-testid="stFileUploader"] section {{
        background: #24242d !important;
        border: 1px solid #30303a !important;
        border-radius: 12px !important;
    }}
    div[data-testid="stFileUploader"] button {{
        color: white !important;
        border-color: #454552 !important;
        background: #111217 !important;
    }}
    .pas-table-wrap {{ max-height: 510px !important; overflow-y: auto !important; overflow-x: auto !important; }}
    .pas-table-wrap thead th {{ position: sticky; top: 0; z-index: 2; }}
    .pas-note, .pas-support, .pas-support * {{ color: #0A0A0A !important; }}
    .pas-sidebar-title {{ color:#fff; font-size:18px; font-weight:950; line-height:1.15; text-align:center; margin: 20px 0 8px; }}
    .pas-yellow-line {{ width:72px; height:4px; background:{PAS_YELLOW}; border-radius:99px; margin: 0 auto 22px; }}
    .pas-sidebar-copy {{ color:#fff !important; font-size:14px; line-height:1.52; font-weight:650; margin-bottom:24px; }}
    .pas-sidebar-rule {{ border-top:1px solid rgba(255,255,255,.22); margin:22px 0; }}
    .pas-sidebar-heading {{ color:{PAS_YELLOW}; font-size:19px; font-weight:950; margin: 0 0 16px; }}
    .pas-nav-row {{ display:grid; grid-template-columns: 26px 1fr; gap:10px; align-items:start; margin: 15px 0; color:#fff; font-weight:750; line-height:1.25; font-size:14px; }}
    .pas-nav-icon svg {{ width:21px; height:21px; stroke:{PAS_YELLOW}; stroke-width:2.4; fill:none; stroke-linecap:round; stroke-linejoin:round; }}
    .pas-sidebar-footer {{ color:#fff; font-size:12px; font-weight:800; margin-top:28px; }}

    .pas-hero {{ display:flex; align-items:center; gap:16px; background: linear-gradient(100deg, #08090b 0%, #151718 70%, #c9aa00 130%) !important; border-radius: 16px !important; padding: 12px 22px !important; margin: 0 0 18px 0 !important; box-shadow: 0 9px 25px rgba(0,0,0,.13) !important; min-height:60px; }}
    .pas-hero-logo {{ width:37px; height:37px; border-radius:7px; background:{PAS_YELLOW}; color:#000; display:inline-flex; align-items:center; justify-content:center; font-weight:950; font-size:14px; letter-spacing:-1px; }}
    .pas-hero-text {{ color:#fff; font-size:18px; font-weight:950; letter-spacing:-.02em; }}
    .pas-hero-dot {{ color:#fff; opacity:.8; margin: 0 7px; }}
    .pas-hero-version {{ color:{PAS_YELLOW}; font-weight:950; }}

    .pas-upload-card {{ background:#fff; border:1px solid #e5e7eb; border-radius:18px; box-shadow:0 5px 18px rgba(15,23,42,.08); padding:16px 18px 14px; margin-bottom:14px; }}
    .pas-upload-title {{ color:#0A0A0A; font-size:16px; font-weight:950; margin-bottom:10px; }}
    div[data-testid="stFileUploader"] {{ margin:0 !important; }}
    div[data-testid="stFileUploader"] label {{ display:none !important; }}
    div[data-testid="stFileUploader"] section {{ background:#f4f6f8 !important; border:1px solid #dfe4ea !important; border-radius:11px !important; min-height:52px !important; padding:8px 10px !important; }}
    div[data-testid="stFileUploader"] section * {{ color:#0A0A0A !important; }}
    div[data-testid="stFileUploader"] button {{ background:#fff !important; color:#0A0A0A !important; border:1px solid #d7dce3 !important; border-radius:10px !important; font-weight:900 !important; box-shadow:0 2px 8px rgba(0,0,0,.06) !important; }}
    div[data-testid="stFileUploader"] svg {{ color:#0A0A0A !important; fill:currentColor !important; stroke:currentColor !important; }}
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] {{ background:#fff !important; border:1px solid #dfe4ea !important; border-radius:10px !important; color:#0A0A0A !important; }}
    div[data-testid="stFileUploader"] small {{ color:#4b5563 !important; }}
    div.stButton > button[kind="secondary"], .stButton > button {{ min-height:52px !important; font-size:16px !important; box-shadow:0 6px 18px rgba(255,212,0,.25) !important; }}
    .stDownloadButton > button {{ min-height:62px !important; font-size:20px !important; box-shadow:0 6px 18px rgba(255,212,0,.25) !important; }}

    .kpi-card {{ background:#fff !important; border-radius:18px !important; border:1px solid #e4e7eb !important; box-shadow:0 5px 20px rgba(15,23,42,.08) !important; min-height:118px !important; padding:18px 22px !important; display:flex; align-items:center; gap:18px; }}
    .kpi-icon {{ width:64px; height:64px; border-radius:50%; background:#fff5bd; display:flex; align-items:center; justify-content:center; flex:none; }}
    .kpi-icon svg {{ width:35px; height:35px; stroke:#0A0A0A; stroke-width:2.5; fill:none; stroke-linecap:round; stroke-linejoin:round; }}
    .kpi-label {{ color:#111 !important; font-size:15px !important; font-weight:950 !important; margin:0 0 3px !important; }}
    .kpi-value {{ color:#e9b900 !important; font-size:42px !important; line-height:.98 !important; font-weight:950 !important; text-shadow:none !important; }}
    .kpi-sub {{ color:#374151 !important; font-size:14px !important; margin-top:6px !important; }}

    .pas-results-title {{ color:#0A0A0A !important; font-size:28px !important; font-weight:950 !important; margin: 22px 0 8px !important; }}
    .pas-unmatched-pill {{ background:{PAS_YELLOW} !important; color:#0A0A0A !important; border:0 !important; border-radius:14px 14px 0 0 !important; padding:13px 20px !important; font-size:18px; box-shadow:0 4px 14px rgba(0,0,0,.09); }}
    .pas-table-wrap {{ background:#fff !important; border:1px solid #e0e4e9 !important; border-radius:0 16px 16px 16px !important; max-height:430px !important; overflow:auto !important; box-shadow:0 7px 25px rgba(15,23,42,.10) !important; }}
    table.pas-table {{ font-size:14px !important; color:#0A0A0A !important; }}
    table.pas-table thead th {{ background:{PAS_YELLOW} !important; color:#0A0A0A !important; border:1px solid #e2ba00 !important; padding:12px 14px !important; font-weight:950 !important; position:sticky; top:0; z-index:5; }}
    table.pas-table tbody td {{ background:#fff !important; color:#0A0A0A !important; border:1px solid #e1e5eb !important; padding:10px 14px !important; }}
    table.pas-table tbody tr:nth-child(even) td {{ background:#fbfcfd !important; }}

    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] {{ display: none !important; }}
    div[data-testid="stFileUploaderDropzone"] {{ background: transparent !important; border: 0 !important; padding: 0 !important; min-height: 0 !important; }}
    div[data-testid="stFileUploaderDropzoneInstructions"] {{ display: none !important; }}
    div[data-testid="stFileUploader"] section {{ background: transparent !important; border: 0 !important; min-height: 0 !important; padding: 0 !important; }}
    .pas-file-card {{ display:flex; align-items:center; gap:14px; background:#f4f6f8; border:1px solid #dfe4ea; border-radius:12px; padding:11px 14px; min-height:54px; margin: 4px 0 12px; }}
    .pas-file-icon {{ width:32px; height:32px; border-radius:8px; display:flex; align-items:center; justify-content:center; color:#fff; font-weight:950; font-size:11px; box-shadow:0 2px 8px rgba(0,0,0,.12); flex:none; }}
    .pas-file-icon.excel {{ background:#118a3b; }}
    .pas-file-main {{ flex:1; min-width:0; }}
    .pas-file-name {{ color:#0A0A0A; font-weight:950; font-size:15px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    .pas-file-size {{ color:#4b5563; font-weight:650; font-size:13px; margin-top:2px; }}
    .pas-file-check {{ width:24px; height:24px; border-radius:50%; background:#108a37; color:white; display:flex; align-items:center; justify-content:center; font-size:15px; font-weight:950; flex:none; }}
    </style>
    """,
    unsafe_allow_html=True,
)



st.markdown(
    """
    <style>
    /* Bottom chase animation: small, low, runs once */
    .pas-bottom-chase-wrap {
        position: fixed;
        left: calc(18rem + 22px);
        right: 42px;
        bottom: 12px;
        height: 58px;
        pointer-events: none;
        z-index: 1;
        overflow: hidden;
    }
    .pas-bottom-ground {
        position: absolute;
        left: 0;
        right: 0;
        bottom: 6px;
        border-bottom: 1px solid rgba(0,0,0,0.11);
    }
    .pas-chase-pack {
        position: absolute;
        bottom: 8px;
        left: -150px;
        width: 150px;
        height: 48px;
        animation: pas-chase-run 13s linear 1 forwards;
    }
    @keyframes pas-chase-run {
        0% { transform: translateX(-120px); opacity: 0; }
        8% { opacity: 1; }
        88% { opacity: 1; }
        100% { transform: translateX(calc(100vw - 90px)); opacity: 0; }
    }
    .pas-truck-mini { position: absolute; left: 0; bottom: 5px; width: 54px; height: 30px; filter: drop-shadow(0 1px 1px rgba(0,0,0,.22)); }
    .pas-truck-bed { position: absolute; left: 0; top: 5px; width: 34px; height: 19px; background: #FFD400; border: 3px solid #0A0A0A; border-radius: 4px 2px 3px 5px; transform: skewX(-10deg); }
    .pas-truck-logo { position: absolute; left: 7px; top: 9px; font-size: 9px; font-weight: 950; color: #0A0A0A; line-height: 1; z-index: 3; }
    .pas-truck-cab { position: absolute; left: 30px; top: 7px; width: 19px; height: 18px; background: #FFD400; border: 3px solid #0A0A0A; border-radius: 3px 5px 3px 2px; z-index: 2; }
    .pas-truck-window { position: absolute; left: 34px; top: 10px; width: 7px; height: 7px; background: #a8d8e8; border: 2px solid #0A0A0A; border-radius: 2px; z-index: 4; }
    .pas-truck-nose { position: absolute; left: 47px; top: 17px; width: 8px; height: 8px; background: #FFD400; border: 3px solid #0A0A0A; border-left: none; border-radius: 0 3px 3px 0; }
    .pas-wheel { position: absolute; bottom: 0; width: 9px; height: 9px; background: #0A0A0A; border: 2px solid #222; border-radius: 50%; animation: pas-wheel-spin .32s linear infinite; z-index: 5; }
    .pas-wheel::after { content: ""; position: absolute; inset: 2px; background: #FFD400; border-radius: 50%; }
    .pas-wheel.back { left: 13px; }
    .pas-wheel.front { left: 41px; }
    @keyframes pas-wheel-spin { to { transform: rotate(360deg); } }
    .pas-speed-lines { position: absolute; left: -30px; top: 17px; width: 24px; height: 18px; }
    .pas-speed-lines span { display:block; height:2px; background:#b9b9b9; margin:4px 0; border-radius:2px; animation: pas-flicker .55s linear infinite; }
    .pas-speed-lines span:nth-child(2) { width: 16px; margin-left: 8px; }
    .pas-speed-lines span:nth-child(3) { width: 11px; margin-left: 13px; }
    @keyframes pas-flicker { 50% { opacity:.25; transform: translateX(-5px); } }
    .pas-dust { position:absolute; left:-5px; bottom:0; width:34px; height:14px; opacity:.75; }
    .pas-dust span { position:absolute; bottom:0; background:#dac6a9; border-radius:50%; animation: pas-dust 1s linear infinite; }
    .pas-dust span:nth-child(1) { width:12px; height:6px; left:0; }
    .pas-dust span:nth-child(2) { width:16px; height:7px; left:10px; animation-delay:.2s; }
    .pas-dust span:nth-child(3) { width:11px; height:5px; left:23px; animation-delay:.4s; }
    @keyframes pas-dust { 50% { transform: translateX(-8px) scale(1.15); opacity:.4; } }
    .pas-stickman { position: absolute; left: 92px; bottom: 5px; width: 28px; height: 34px; animation: pas-runner-bob .35s ease-in-out infinite alternate; }
    @keyframes pas-runner-bob { from { transform: translateY(1px); } to { transform: translateY(-2px); } }
    .pas-stick-head { position:absolute; top:0; left:11px; width:8px; height:8px; border:2px solid #111; border-radius:50%; background:white; }
    .pas-stick-body { position:absolute; left:15px; top:9px; width:2px; height:13px; background:#111; transform: rotate(12deg); transform-origin:top; }
    .pas-stick-arm-a, .pas-stick-arm-b, .pas-stick-leg-a, .pas-stick-leg-b { position:absolute; width:2px; height:12px; background:#111; transform-origin:top; border-radius:2px; }
    .pas-stick-arm-a { left:15px; top:11px; transform: rotate(58deg); animation: pas-arm-a .35s linear infinite alternate; }
    .pas-stick-arm-b { left:15px; top:11px; transform: rotate(-50deg); animation: pas-arm-b .35s linear infinite alternate; }
    .pas-stick-leg-a { left:16px; top:21px; height:14px; transform: rotate(48deg); animation: pas-leg-a .35s linear infinite alternate; }
    .pas-stick-leg-b { left:16px; top:21px; height:14px; transform: rotate(-42deg); animation: pas-leg-b .35s linear infinite alternate; }
    @keyframes pas-arm-a { to { transform: rotate(-45deg); } }
    @keyframes pas-arm-b { to { transform: rotate(55deg); } }
    @keyframes pas-leg-a { to { transform: rotate(-45deg); } }
    @keyframes pas-leg-b { to { transform: rotate(48deg); } }
    </style>
    """,
    unsafe_allow_html=True,
)


def render_bottom_chase():
    st.markdown(
        """
        <div class="pas-bottom-chase-wrap" aria-hidden="true">
            <div class="pas-bottom-ground"></div>
            <div class="pas-chase-pack">
                <div class="pas-speed-lines"><span></span><span></span><span></span></div>
                <div class="pas-dust"><span></span><span></span><span></span></div>
                <div class="pas-truck-mini">
                    <div class="pas-truck-bed"></div>
                    <div class="pas-truck-logo">PAS</div>
                    <div class="pas-truck-cab"></div>
                    <div class="pas-truck-window"></div>
                    <div class="pas-truck-nose"></div>
                    <div class="pas-wheel back"></div>
                    <div class="pas-wheel front"></div>
                </div>
                <div class="pas-stickman">
                    <div class="pas-stick-head"></div>
                    <div class="pas-stick-body"></div>
                    <div class="pas-stick-arm-a"></div>
                    <div class="pas-stick-arm-b"></div>
                    <div class="pas-stick-leg-a"></div>
                    <div class="pas-stick-leg-b"></div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def logo_available() -> bool:
    return Path("pas_logo.png").exists() or Path("assets/pas_logo.png").exists()

with st.sidebar:
    if Path("pas_logo.png").exists():
        st.image("pas_logo.png", use_column_width=True)
    elif Path("assets/pas_logo.png").exists():
        st.image("assets/pas_logo.png", use_column_width=True)
    else:
        st.markdown('<div style="background:#FFD400;color:#000;border-radius:14px;padding:18px;text-align:center;font-weight:950;font-size:30px;">PAS</div>', unsafe_allow_html=True)
    st.markdown(
        """
        <div class="pas-sidebar-title">PAS Vendor<br>On-Hire Checker</div>
        <div class="pas-yellow-line"></div>
        <div class="pas-sidebar-copy">Upload vendor hire report and Materials & Plant Orders, then export matched/unmatched Excel.</div>
        <div class="pas-sidebar-rule"></div>
        <div class="pas-sidebar-heading">Instructions</div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M16 16l-4-4-4 4"/><path d="M12 12v9"/><path d="M20 16.6A5 5 0 0 0 18 7h-1.3A8 8 0 1 0 4 15.3"/></svg></span><span>Upload Vendor Hire Report</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M5 3l14 9-14 9V3z"/></svg></span><span>Run On-Hire Check</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><path d="M7 10l5 5 5-5"/><path d="M12 15V3"/></svg></span><span>Download Excel</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><path d="M7 10l5 5 5-5"/><path d="M12 15V3"/></svg></span><span>Matched / Unmatched Only</span></div>
        <div class="pas-nav-row"><span class="pas-nav-icon"><svg viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="M21 21l-4.3-4.3"/></svg></span><span>Smoke Crack</span></div>
        <div class="pas-sidebar-rule"></div>
        <div class="pas-sidebar-footer">PAS NW Ltd • v1.0 Prototype Build</div>
        """,
        unsafe_allow_html=True,
    )

st.markdown(
    f"""
    <div class="pas-hero">
      <div class="pas-hero-logo">PAS</div>
      <div class="pas-hero-text">PAS NW Ltd<span class="pas-hero-dot">•</span><span class="pas-hero-version">{APP_VERSION}</span></div>
    </div>
    """,
    unsafe_allow_html=True,
)


def render_selected_file_card(uploaded_file, file_kind="excel"):
    size = getattr(uploaded_file, "size", 0) or 0
    if size >= 1024 * 1024:
        size_text = f"{size / (1024 * 1024):.1f} MB"
    else:
        size_text = f"{size / 1024:.0f} KB"
    st.markdown(
        f'''
        <div class="pas-file-card">
            <div class="pas-file-icon excel">XLS</div>
            <div class="pas-file-main">
                <div class="pas-file-name">{escape(getattr(uploaded_file, "name", "Uploaded file"))}</div>
                <div class="pas-file-size">{size_text}</div>
            </div>
            <div class="pas-file-check">✓</div>
        </div>
        ''',
        unsafe_allow_html=True,
    )



# ===== Vendor On-Hire Checker logic =====
from difflib import SequenceMatcher
from typing import Optional, List, Tuple, Dict

GREEN_FILL = "C6EFCE"
RED_FILL = "FFC7CE"
ORANGE_FILL = "F4B183"
HEADER_FILL = "FFD400"
WHITE_FILL = "FFFFFF"


def clean_cell(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def norm_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_cell(value).lower())


def find_col(columns: List[str], aliases: List[str]) -> Optional[str]:
    normed = {c: norm_header(c) for c in columns}
    alias_norms = [norm_header(a) for a in aliases]
    for alias in alias_norms:
        for col, ncol in normed.items():
            if alias == ncol:
                return col
    for alias in alias_norms:
        for col, ncol in normed.items():
            if alias and alias in ncol:
                return col
    return None


def money_to_float(value) -> Optional[float]:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = clean_cell(value).replace(",", "")
    if not text:
        return None
    matches = re.findall(r"-?£?\s*([0-9]+(?:\.[0-9]+)?)", text)
    if not matches:
        return None
    try:
        return round(float(matches[-1]), 2)
    except Exception:
        return None


def normalise_text(value) -> str:
    text = clean_cell(value).lower()
    replacements = {
        "3 tonne": "3t", "3 ton": "3t", "3-ton": "3t", "3 tonne": "3t",
        "1 tonne": "1t", "1 ton": "1t", "1-ton": "1t",
        "5 tonne": "5t", "5 ton": "5t", "5-ton": "5t",
        "digger": "excavator", "mini digger": "excavator", "exc": "excavator",
        "wacker plate": "plate compactor", "vib plate": "plate compactor", "vibrating plate": "plate compactor",
        "stihl saw": "cut off saw", "stihl": "cut off saw", "disc cutter": "cut off saw",
        "telehandler": "telehandler", "fork lift": "forklift", "fork-lift": "forklift",
        "welfare cabin": "welfare unit", "site cabin": "welfare unit", "cabin": "welfare unit",
        "10ft steel store": "10ft container", "20ft steel store": "20ft container",
        "steel store": "container", "site store": "container", "storage container": "container",
        "shipping container": "container", "store unit": "container",
        "gen set": "generator", "genny": "generator",
        "qty": "quantity",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    stop = {
        "the", "and", "with", "for", "hire", "weekly", "week", "charge", "item", "plant", "no",
        "serial", "fleet", "make", "model", "pas", "nw", "ltd", "each", "day", "days", "wk"
    }
    tokens = [t for t in text.split() if t not in stop]
    return " ".join(tokens)


def similarity(a, b) -> float:
    a2, b2 = normalise_text(a), normalise_text(b)
    if not a2 or not b2:
        return 0.0
    set_a, set_b = set(a2.split()), set(b2.split())
    if not set_a or not set_b:
        return SequenceMatcher(None, a2, b2).ratio()
    token_overlap = len(set_a & set_b) / max(1, min(len(set_a), len(set_b)))
    seq = SequenceMatcher(None, a2, b2).ratio()
    contained = 1.0 if a2 in b2 or b2 in a2 else 0.0
    return max(seq, token_overlap, contained)


def extract_job(value) -> str:
    text = clean_cell(value).upper()
    m = re.search(r"\b(P\d{2,4})(?:[A-Z0-9&/-]*)?\b", text)
    return m.group(1) if m else ""


def extract_hire_no(value) -> str:
    text = clean_cell(value)
    h = extract_hire_ref(text)
    if h:
        return h[1:]
    m = re.search(r"\b(\d{4,8})\b", text)
    return m.group(1) if m else ""


def extract_hire_ref(value) -> str:
    """Return H reference such as H6020 from an order/reference string."""
    text = clean_cell(value).upper()
    m = re.search(r"\bH\s*[-/]?\s*(\d{3,8})\b", text)
    return f"H{m.group(1)}" if m else ""


def normalise_fleet(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_cell(value).lower())


def is_live_status(status: str) -> bool:
    s = clean_cell(status).lower().strip()
    return s in {"on hire", "missing"}


def is_off_hired_status(status: str) -> bool:
    s = clean_cell(status).lower().strip()
    return "off" in s and "hire" in s


def is_operated_plant_text(*values) -> bool:
    blob = " ".join(clean_cell(v).lower() for v in values)
    return "operated plant" in blob or "operated" in blob and "plant" in blob


def parse_any_date(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_cell(value)
    if not text:
        return None
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def pas_item_live_as_of(prow, as_of_date=None) -> bool:
    if as_of_date is None:
        as_of_date = datetime.now().date()
    status = clean_cell(prow.get("PAS Status", "")).lower().strip()
    if is_live_status(status):
        return True
    if is_off_hired_status(status):
        off_hire_date = parse_any_date(prow.get("PAS Off Hire Date", ""))
        if off_hire_date and off_hire_date >= as_of_date:
            return True
    return False


def off_hire_reason(prow, as_of_date=None) -> str:
    if as_of_date is None:
        as_of_date = datetime.now().date()
    off_hire_date = parse_any_date(prow.get("PAS Off Hire Date", ""))
    if off_hire_date:
        return f"PAS item is off-hired before check date ({off_hire_date.strftime('%d/%m/%Y')})"
    return "PAS item is off-hired"


def read_excel_any(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".xls"):
        return pd.read_excel(uploaded_file, engine="xlrd")
    return pd.read_excel(uploaded_file)




def parse_vendor_live_hire_pdf(uploaded_file) -> pd.DataFrame:
    """Convert supplier live-hire PDFs into clean line-level rows before matching.

    Handles both known supplier styles used so far:
    1) Ambrose-style live hire detail reports with Order No blocks and weekly/net rates.
    2) Table-style equipment-on-hire PDFs with Site / Cat / Stock No / Qty / Description / Date / Contract / Order.

    The function returns the same normalised vendor schema used by Excel imports.
    """
    try:
        from pypdf import PdfReader
    except Exception:
        try:
            from PyPDF2 import PdfReader
        except Exception:
            raise RuntimeError("PDF reader unavailable. Add pypdf to requirements.txt.")

    pdf_bytes = uploaded_file.read()
    uploaded_file.seek(0)
    reader = PdfReader(io.BytesIO(pdf_bytes))

    all_lines = []
    for page_no, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        for raw_line in text.splitlines():
            line = re.sub(r"\s+", " ", raw_line).strip()
            if line:
                all_lines.append((page_no, line))

    rows = []

    # ------------------------------------------------------------------
    # Parser A: table-style equipment-on-hire PDFs.
    # Example:
    # Site: PEEL HALL, WARRINGTON, WA2 9UF
    # BAT2 HNB01133 1.00 HILTI NURON B 22-110 22V BATTERY 05-Jan-26 008-703731 P151/H7044
    # ------------------------------------------------------------------
    current_site = ""
    table_item_re = re.compile(
        r'^(?P<cat>[A-Z0-9][A-Z0-9/-]{1,20})\s+'
        r'(?P<stock>[A-Z0-9][A-Z0-9/-]{1,30})\s+'
        r'(?P<qty>-?\d+(?:\.\d+)?)\s+'
        r'(?P<desc>.+?)\s+'
        r'(?P<onhire>\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4}|\d{1,2}/\d{1,2}/\d{2,4})\s+'
        r'(?P<contract>[A-Za-z0-9-]+)\s+'
        r'(?P<order>[A-Za-z0-9/&.-]+)\s*$',
        re.I,
    )
    table_header_seen = False
    for page_no, line in all_lines:
        low = line.lower()
        if low.startswith("site:"):
            current_site = line.split(":", 1)[1].strip()
            continue
        if "cat stock no qty description date on hire contract no order no" in low:
            table_header_seen = True
            continue
        if not table_header_seen:
            continue
        if low.startswith("equipment currently on hire") or low.startswith("cat stock no"):
            continue
        m = table_item_re.match(line)
        if not m:
            continue
        rows.append({
            "Vendor Site": current_site,
            "Vendor Fleet No": m.group("stock").strip(),
            "Vendor Description": m.group("desc").strip(),
            "Vendor Qty": m.group("qty"),
            "Vendor On Hire Date": m.group("onhire"),
            "Vendor Contract No": m.group("contract"),
            "Vendor Order No": m.group("order"),
            "Vendor Rate": "",
            "Vendor Weekly Rate": "",
            "Vendor Net Weekly": "",
            "Vendor Last Inv Date": "",
            "Vendor Type": "Hire",
            "Vendor Cat": m.group("cat").strip(),
            "Source Page": page_no,
        })

    # ------------------------------------------------------------------
    # Parser B: Ambrose-style live hire detail reports.
    # Only run if the table parser did not find anything.
    # ------------------------------------------------------------------
    if not rows:
        current_contract = ""
        current_order = ""
        current_site = ""
        pending_address_parts = []

        item_re = re.compile(
            r'^(?P<item>[A-Z0-9][A-Z0-9/-]{1,20})\s+'
            r'(?P<desc>.+?)\s+'
            r'(?P<qty>-?\d+(?:\.\d+)?)\s+'
            r'(?P<weekly>-?\d+(?:\.\d+)?)\s+'
            r'(?P<discount>-?\d+(?:\.\d+)?%)\s+'
            r'(?P<net>-?\d+(?:\.\d+)?)\s+'
            r'(?P<lastinv>N/A|\d{1,2}/\d{1,2}/\d{2,4})\s+'
            r'(?P<onhire>\d{1,2}/\d{1,2}/\d{2,4})'
            r'(?:\s+\d{1,2}:\d{2}:\d{2})?\s+'
            r'(?P<type>[A-Za-z]+)\s*$',
            re.I,
        )
        contract_re = re.compile(
            r'^(?P<contract>\d{6,12})\s+.*?(?:\bOrder\s+No:\s*(?P<order_after>[A-Za-z0-9/&.-]+)|(?P<order_before>[A-Za-z0-9]+/H\d+)\s*Order\s+No:)(?P<tail>.*)$',
            re.I,
        )
        page_footer_re = re.compile(r'^\d{1,2}/\d{1,2}/\d{2,4}\s+Page\s+\d+\s+of\s+\d+', re.I)

        def commit_site_from_pending():
            nonlocal current_site, pending_address_parts
            if pending_address_parts:
                current_site = " ".join(pending_address_parts).strip(" ,")
                pending_address_parts = []

        for page_no, line in all_lines:
            low = line.lower()
            if (
                "live hire detail report" in low
                or line in {"PAS002", "Type"}
                or low.startswith("customer range")
                or low.startswith("item no description")
                or low.startswith("contract no acct no")
                or low.startswith("depot:")
                or page_footer_re.match(line)
            ):
                continue

            cm = contract_re.match(line)
            if cm:
                commit_site_from_pending()
                current_contract = cm.group("contract")
                current_order = cm.group("order_after") or cm.group("order_before") or ""
                current_site = ""
                pending_address_parts = []
                tail = cm.group("tail") or ""
                if "Delivery Address" in tail:
                    address = tail.split("Delivery Address", 1)[1].strip(" :,-")
                    if address:
                        pending_address_parts.append(address)
                continue

            im = item_re.match(line)
            if im:
                commit_site_from_pending()
                rows.append({
                    "Vendor Site": current_site,
                    "Vendor Fleet No": im.group("item").strip(),
                    "Vendor Description": im.group("desc").strip(),
                    "Vendor Qty": im.group("qty"),
                    "Vendor On Hire Date": im.group("onhire"),
                    "Vendor Contract No": current_contract,
                    "Vendor Order No": current_order,
                    "Vendor Rate": im.group("net"),
                    "Vendor Weekly Rate": im.group("weekly"),
                    "Vendor Net Weekly": im.group("net"),
                    "Vendor Last Inv Date": im.group("lastinv"),
                    "Vendor Type": im.group("type"),
                    "Source Page": page_no,
                })
                continue

            if current_order and not re.match(r'^\d{6,12}\s+', line):
                if not item_re.match(line):
                    pending_address_parts.append(line)

    if not rows:
        raise RuntimeError("Could not convert the PDF into hire lines. This PDF layout is not recognised yet.")

    out = pd.DataFrame(rows)
    for col in [
        "Vendor Site", "Vendor Fleet No", "Vendor Description", "Vendor Qty", "Vendor On Hire Date",
        "Vendor Contract No", "Vendor Order No", "Vendor Rate", "Vendor Weekly Rate", "Vendor Net Weekly",
        "Vendor Last Inv Date", "Vendor Type", "Source Page"
    ]:
        if col not in out.columns:
            out[col] = ""
    out["Vendor Rate Value"] = out["Vendor Rate"].apply(money_to_float)
    out["Vendor Job"] = out["Vendor Order No"].apply(extract_job)
    out["Vendor Hire No"] = out["Vendor Order No"].apply(extract_hire_no)
    out["Vendor Hire Ref"] = out["Vendor Order No"].apply(extract_hire_ref)
    out["Vendor Row No"] = range(2, len(out) + 2)
    return out

def load_vendor_report(uploaded_file) -> Tuple[pd.DataFrame, pd.DataFrame]:
    name = uploaded_file.name.lower()
    if name.endswith(".pdf"):
        out = parse_vendor_live_hire_pdf(uploaded_file)
        raw = out.copy()
        return raw, out

    raw = read_excel_any(uploaded_file)
    raw = raw.dropna(how="all").copy()
    cols = list(raw.columns)
    mapping = {
        "Vendor Site": find_col(cols, ["Site", "Location", "Project", "Job", "Site Name"]),
        "Vendor Fleet No": find_col(cols, ["Item No", "Fleet No", "Fleet", "Asset", "Serial", "Plant No", "Registration"]),
        "Vendor Description": find_col(cols, ["Description", "Item Description", "Product", "Equipment", "Plant Description", "Name"]),
        "Vendor Qty": find_col(cols, ["Quantity", "Qty"]),
        "Vendor On Hire Date": find_col(cols, ["Date", "On Hire Date", "Start Date", "Delivery Date", "Hired Date"]),
        "Vendor Contract No": find_col(cols, ["Syrinx Contract No", "Contract No", "Contract", "Hire Contract"]),
        "Vendor Order No": find_col(cols, ["Order No", "PO", "Purchase Order", "Order Number", "Customer Order"]),
        "Vendor Rate": find_col(cols, ["Hire Rate", "Rate", "Weekly Rate", "Cost", "Value", "Charge", "Amount", "Net Weekly"]),
    }
    out = pd.DataFrame()
    for new_col, old_col in mapping.items():
        out[new_col] = raw[old_col] if old_col in raw.columns else ""
    if "Vendor Description" not in out or out["Vendor Description"].replace("", pd.NA).dropna().empty:
        # fallback: first meaningful text column
        for col in cols:
            if raw[col].dtype == object:
                out["Vendor Description"] = raw[col]
                break
    out["Vendor Rate Value"] = out.get("Vendor Rate", "").apply(money_to_float) if isinstance(out.get("Vendor Rate", ""), pd.Series) else None
    out["Vendor Job"] = out.get("Vendor Order No", "").apply(extract_job) if isinstance(out.get("Vendor Order No", ""), pd.Series) else ""
    out["Vendor Hire No"] = out.get("Vendor Order No", "").apply(extract_hire_no) if isinstance(out.get("Vendor Order No", ""), pd.Series) else ""
    out["Vendor Row No"] = range(2, len(out) + 2)
    return raw, out


def load_pas_plant(uploaded_file) -> pd.DataFrame:
    xls = pd.ExcelFile(uploaded_file)
    sheet = "Plant" if "Plant" in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(xls, sheet_name=sheet).dropna(how="all").copy()
    cols = list(df.columns)
    def pick(names):
        col = find_col(cols, names)
        return df[col] if col in df.columns else ""
    out = pd.DataFrame()
    out["PAS Description"] = pick(["Description", "Item Description", "Plant Description"])
    out["PAS Fleet No"] = pick(["Fleet No.", "Fleet No", "Fleet", "Item No", "Serial"])
    out["PAS Supplier"] = pick(["Supplier", "Vendor"])
    out["PAS Qty"] = pick(["Qty", "Quantity"])
    out["PAS On Hire Date"] = pick(["On Hire / Delivery Date", "On Hire Date", "Delivery Date"])
    out["PAS Expected Off Hire Date"] = pick(["Expected Off-Hire Date", "Expected Off Hire Date"])
    out["PAS Off Hire Date"] = pick(["Off Hire Date", "Off-Hire Date"])
    out["PAS Status"] = pick(["Status"])
    out["PAS Job No"] = pick(["Job No", "Job Number", "Job"])
    out["PAS Order Number"] = pick(["Order Number", "Sage Order No", "PO", "Purchase Order"])
    out["PAS Site Name"] = pick(["Site Name", "Site"])
    out["PAS Row No"] = range(2, len(out) + 2)
    out["PAS Job Base"] = out["PAS Job No"].apply(extract_job)
    out["PAS Hire Ref"] = out["PAS Order Number"].apply(extract_hire_ref)
    out["PAS Hire No"] = out["PAS Hire Ref"].apply(lambda x: x[1:] if x else "")
    return out


def vendor_zero_cost(vrow) -> bool:
    val = vrow.get("Vendor Rate Value", None)
    return val is not None and not pd.isna(val) and abs(float(val)) < 0.0001


def vendor_match_description(vrow) -> str:
    group_desc = clean_cell(vrow.get("Vendor Group Description", ""))
    return group_desc or clean_cell(vrow.get("Vendor Description", ""))


def build_vendor_groups(vendor_df: pd.DataFrame) -> pd.DataFrame:
    """Group related vendor lines before matching.

    This keeps lines such as "10ft Steel Store" + "Padlock" together when they
    share the same order/contract, so they can match one PAS line like
    "Container: 10ft & Padlock".
    """
    if vendor_df.empty:
        return vendor_df
    df = vendor_df.copy()
    for col in ["Vendor Order No", "Vendor Contract No", "Vendor Description"]:
        if col not in df.columns:
            df[col] = ""

    def group_key(row):
        order = clean_cell(row.get("Vendor Order No", "")).upper()
        contract = clean_cell(row.get("Vendor Contract No", "")).upper()
        if order or contract:
            return f"{order}|{contract}"
        return f"ROW|{row.name}"

    df["_Vendor Group Key"] = df.apply(group_key, axis=1)
    group_desc_map = {}
    group_count_map = {}
    for key, grp in df.groupby("_Vendor Group Key", sort=False):
        descriptions = []
        for value in grp["Vendor Description"].tolist():
            cleaned = clean_cell(value)
            if cleaned and cleaned not in descriptions:
                descriptions.append(cleaned)
        group_desc_map[key] = " + ".join(descriptions)
        group_count_map[key] = len(grp)
    df["Vendor Group Description"] = df["_Vendor Group Key"].map(group_desc_map)
    df["Vendor Group Line Count"] = df["_Vendor Group Key"].map(group_count_map)
    return df


def score_candidate(vrow, prow) -> Tuple[float, List[str]]:
    reasons = []
    match_desc = vendor_match_description(vrow)
    score = similarity(match_desc, prow.get("PAS Description", "")) * 75

    vhire = clean_cell(vrow.get("Vendor Hire Ref", "")) or ("H" + clean_cell(vrow.get("Vendor Hire No", "")) if clean_cell(vrow.get("Vendor Hire No", "")) else "")
    phire = clean_cell(prow.get("PAS Hire Ref", "")) or ("H" + clean_cell(prow.get("PAS Hire No", "")) if clean_cell(prow.get("PAS Hire No", "")) else "")
    if vhire and phire and vhire.upper() == phire.upper():
        score += 45
        reasons.append("H number matched")

    vjob = clean_cell(vrow.get("Vendor Job", ""))
    pjob = clean_cell(prow.get("PAS Job Base", "")) or extract_job(prow.get("PAS Job No", ""))
    if vjob and pjob and vjob.upper() == pjob.upper():
        score += 20
        reasons.append("Job/PO base matched")
    elif vjob and pjob and vjob.upper() != pjob.upper():
        score -= 100
        reasons.append("Job/PO base rejected")

    vfleet = normalise_fleet(vrow.get("Vendor Fleet No", ""))
    pfleet = normalise_fleet(prow.get("PAS Fleet No", ""))
    if vfleet and pfleet and vfleet == pfleet:
        score += 10
        reasons.append("Fleet matched")
    return score, reasons


def find_best_match(vrow, pas_df: pd.DataFrame) -> Tuple[Optional[pd.Series], float, str]:
    if pas_df.empty:
        return None, 0.0, "No PAS rows loaded"

    candidates = pas_df.copy()
    vhire_ref = clean_cell(vrow.get("Vendor Hire Ref", ""))
    if not vhire_ref and clean_cell(vrow.get("Vendor Hire No", "")):
        vhire_ref = "H" + clean_cell(vrow.get("Vendor Hire No", ""))
    vjob = clean_cell(vrow.get("Vendor Job", ""))

    # Hard gate 1: if the vendor gives an H number, never match against a different H number.
    if vhire_ref:
        hire_matches = candidates[candidates.get("PAS Hire Ref", pd.Series(dtype=str)).fillna("").astype(str).str.upper().eq(vhire_ref.upper())]
        if hire_matches.empty:
            return None, 0.0, f"No PAS row found with {vhire_ref}"
        candidates = hire_matches
    # Hard gate 2: if no H number is available but a P/job base is available, stay inside that job.
    elif vjob:
        job_matches = candidates[candidates["PAS Job Base"].fillna("").astype(str).str.upper().eq(vjob.upper())]
        if job_matches.empty:
            return None, 0.0, f"No PAS row found for {vjob}"
        candidates = job_matches

    best = None
    best_score = -999.0
    best_reason = ""
    for _, prow in candidates.iterrows():
        score, reasons = score_candidate(vrow, prow)
        if score > best_score:
            best = prow
            best_score = score
            best_reason = "; ".join(reasons) or "Description compared"
    return best, best_score, best_reason


def reconcile(vendor_df: pd.DataFrame, pas_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    checked_rows = []
    ignored_rows = []
    as_of_date = datetime.now().date()
    vendor_df = build_vendor_groups(vendor_df)

    for _, vrow in vendor_df.iterrows():
        if vendor_zero_cost(vrow):
            ignored_rows.append({**vrow.to_dict(), "Ignored Reason": "£0 vendor line ignored"})
            continue

        best, score, reason = find_best_match(vrow, pas_df)
        status = "Unmatched"
        result_reason = reason or "No live PAS item found"
        fleet_mismatch = False

        if best is not None:
            match_desc = vendor_match_description(vrow)
            desc_score = similarity(match_desc, best.get("PAS Description", ""))
            desc_ok = desc_score >= 0.25

            vjob = clean_cell(vrow.get("Vendor Job", ""))
            pjob = clean_cell(best.get("PAS Job Base", "")) or extract_job(best.get("PAS Job No", ""))
            job_ok = True
            if vjob and pjob:
                job_ok = vjob.upper() == pjob.upper()

            vhire = clean_cell(vrow.get("Vendor Hire Ref", "")) or ("H" + clean_cell(vrow.get("Vendor Hire No", "")) if clean_cell(vrow.get("Vendor Hire No", "")) else "")
            phire = clean_cell(best.get("PAS Hire Ref", "")) or ("H" + clean_cell(best.get("PAS Hire No", "")) if clean_cell(best.get("PAS Hire No", "")) else "")
            hire_ok = True
            if vhire and phire:
                hire_ok = vhire.upper() == phire.upper()

            # If the H number or P/job is right, allow vague/grouped descriptions.
            # This covers accessories/package lines such as Steel Store + Padlock -> Container & Padlock.
            gated_match = bool((vhire and phire and hire_ok) or (vjob and pjob and job_ok))

            if is_operated_plant_text(vrow.get("Vendor Description", ""), match_desc, best.get("PAS Description", ""), best.get("PAS Status", "")):
                status = "Unmatched"
                result_reason = "Operated plant"
            elif not hire_ok or not job_ok:
                status = "Unmatched"
                result_reason = "Wrong/missing PO with no sensible match"
            elif not pas_item_live_as_of(best, as_of_date):
                if is_off_hired_status(best.get("PAS Status", "")):
                    result_reason = off_hire_reason(best, as_of_date)
                else:
                    result_reason = "No live PAS item found"
            elif score < 20 and not desc_ok and not gated_match:
                status = "Unmatched"
                result_reason = "No live PAS item found"
            else:
                status = "Matched"
                result_reason = "Live PAS item found"
                vfleet = normalise_fleet(vrow.get("Vendor Fleet No", ""))
                pfleet = normalise_fleet(best.get("PAS Fleet No", ""))
                fleet_mismatch = bool(vfleet and pfleet and vfleet != pfleet)

        record = {
            "Status": status,
            "Reason": result_reason,
            "Fleet Mismatch": "Yes" if fleet_mismatch and status == "Matched" else "",
            "Match Score": round(float(score), 1),
            "Vendor Site": vrow.get("Vendor Site", ""),
            "Vendor Order No": vrow.get("Vendor Order No", ""),
            "Vendor Contract No": vrow.get("Vendor Contract No", ""),
            "Vendor Fleet No": vrow.get("Vendor Fleet No", ""),
            "Vendor Description": vrow.get("Vendor Description", ""),
            "Vendor Group Description": vrow.get("Vendor Group Description", ""),
            "Vendor Group Line Count": vrow.get("Vendor Group Line Count", ""),
            "Vendor Qty": vrow.get("Vendor Qty", ""),
            "Vendor On Hire Date": vrow.get("Vendor On Hire Date", ""),
            "PAS Job No": best.get("PAS Job No", "") if best is not None else "",
            "PAS Site Name": best.get("PAS Site Name", "") if best is not None else "",
            "PAS Status": best.get("PAS Status", "") if best is not None else "",
            "PAS Fleet No": best.get("PAS Fleet No", "") if best is not None else "",
            "PAS Description": best.get("PAS Description", "") if best is not None else "",
            "PAS On Hire Date": best.get("PAS On Hire Date", "") if best is not None else "",
            "PAS Off Hire Date": best.get("PAS Off Hire Date", "") if best is not None else "",
        }
        checked_rows.append(record)
    return pd.DataFrame(checked_rows), pd.DataFrame(ignored_rows)


def autosize_ws(ws):
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            max_len = max(max_len, len(clean_cell(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 55)


def make_output_excel(checked_df: pd.DataFrame, ignored_df: pd.DataFrame, summary_df: pd.DataFrame) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        checked_df.to_excel(writer, index=False, sheet_name="Vendor Report Checked")
        checked_df[checked_df["Status"] == "Matched"].to_excel(writer, index=False, sheet_name="Matched")
        checked_df[checked_df["Status"] == "Unmatched"].to_excel(writer, index=False, sheet_name="Unmatched")
        if not ignored_df.empty:
            ignored_df.to_excel(writer, index=False, sheet_name="Ignored £0 Items")
        wb = writer.book
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor=HEADER_FILL)
        green = PatternFill("solid", fgColor=GREEN_FILL)
        red = PatternFill("solid", fgColor=RED_FILL)
        orange = PatternFill("solid", fgColor=ORANGE_FILL)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 25
            ws.auto_filter.ref = ws.dimensions
            headers = [clean_cell(c.value) for c in ws[1]]
            status_col = headers.index("Status") + 1 if "Status" in headers else None
            vendor_fleet_col = headers.index("Vendor Fleet No") + 1 if "Vendor Fleet No" in headers else None
            fleet_mismatch_col = headers.index("Fleet Mismatch") + 1 if "Fleet Mismatch" in headers else None
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(name="Calibri", size=10, bold=(cell.row == 1), color="000000")
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = border
                    if cell.row == 1:
                        cell.fill = header_fill
                if cell.row == 1:
                    continue
                if status_col:
                    status = clean_cell(ws.cell(cell.row, status_col).value)
                    fill = green if status == "Matched" else red if status == "Unmatched" else None
                    if fill:
                        for c in range(1, ws.max_column + 1):
                            ws.cell(cell.row, c).fill = fill
                    if vendor_fleet_col and fleet_mismatch_col and clean_cell(ws.cell(cell.row, fleet_mismatch_col).value) == "Yes":
                        ws.cell(cell.row, vendor_fleet_col).fill = orange
            autosize_ws(ws)
    return out.getvalue()


def render_results_table(df: pd.DataFrame):
    if df is None or df.empty:
        st.markdown('<div class="pas-unmatched-pill">Unmatched Items</div>', unsafe_allow_html=True)
        st.markdown('<div class="pas-table-wrap"><table class="pas-table"><tbody><tr><td>No unmatched items found.</td></tr></tbody></table></div>', unsafe_allow_html=True)
        return
    display_cols = ["Reason", "Vendor Order No", "Vendor Fleet No", "Vendor Description", "PAS Status", "PAS Job No", "PAS Description"]
    display_df = df[df["Status"] == "Unmatched"].copy()
    if display_df.empty:
        st.markdown('<div class="pas-unmatched-pill">Unmatched Items</div>', unsafe_allow_html=True)
        st.markdown('<div class="pas-table-wrap"><table class="pas-table"><tbody><tr><td>No unmatched items found.</td></tr></tbody></table></div>', unsafe_allow_html=True)
        return
    display_df = display_df[[c for c in display_cols if c in display_df.columns]].head(200)
    header_html = "".join(f"<th>{escape(c)}</th>" for c in display_df.columns)
    rows_html = []
    for _, row in display_df.iterrows():
        rows_html.append("<tr>" + "".join(f"<td>{escape(clean_cell(row.get(c, '')))}</td>" for c in display_df.columns) + "</tr>")
    st.markdown('<div class="pas-unmatched-pill">Unmatched Items</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="pas-table-wrap"><table class="pas-table"><thead><tr>{header_html}</tr></thead><tbody>{"".join(rows_html)}</tbody></table></div>', unsafe_allow_html=True)


# ===== UI: same uploaded layout, vendor checker content only =====
up_col1, up_col2 = st.columns(2)
with up_col1:
    st.markdown('<div class="pas-upload-card"><div class="pas-upload-title">Upload Vendor Hire Report</div>', unsafe_allow_html=True)
    vendor_file = st.file_uploader("Upload Vendor Hire Report", type=["xlsx", "xls", "pdf"], label_visibility="collapsed", key="vendor_upload")
    if vendor_file:
        render_selected_file_card(vendor_file, "excel")
    st.markdown('</div>', unsafe_allow_html=True)
with up_col2:
    st.markdown('<div class="pas-upload-card"><div class="pas-upload-title">Upload Material & Plant Orders workbook</div>', unsafe_allow_html=True)
    orders_file = st.file_uploader("Upload Material & Plant Orders", type=["xlsx", "xlsm", "xls"], label_visibility="collapsed", key="orders_upload")
    if orders_file:
        render_selected_file_card(orders_file, "excel")
    st.markdown('</div>', unsafe_allow_html=True)

if vendor_file and orders_file:
    st.markdown(
        """
        <style>
        /* Hide Streamlit's black uploaded-file chips once both files are selected.
           Keep the PAS file cards visible. */
        div[data-testid="stFileUploader"] {
            display: none !important;
            visibility: hidden !important;
            height: 0 !important;
            min-height: 0 !important;
            margin: 0 !important;
            padding: 0 !important;
            overflow: hidden !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

run = st.button("▶  Run on-hire check", use_container_width=True)

if "vendor_on_hire_results" not in st.session_state:
    st.session_state["vendor_on_hire_results"] = None

if run:
    if not vendor_file or not orders_file:
        st.warning("Please upload both the vendor hire report and the Material & Plant Orders workbook.")
        st.stop()
    try:
        with st.spinner("Reading vendor hire report..."):
            raw_vendor, vendor_df = load_vendor_report(vendor_file)
        with st.spinner("Reading PAS Material & Plant Orders..."):
            pas_df = load_pas_plant(orders_file)
        with st.spinner("Checking whether vendor items are still live on PAS report..."):
            checked_df, ignored_df = reconcile(vendor_df, pas_df)
        if checked_df.empty:
            st.warning("No chargeable vendor lines were found after ignoring £0 lines.")
            st.stop()
        total = len(checked_df)
        matched = int((checked_df["Status"] == "Matched").sum())
        unmatched = int((checked_df["Status"] == "Unmatched").sum())
        ignored = len(ignored_df)
        match_pct = round((matched / total) * 100, 1) if total else 0.0
        summary_df = pd.DataFrame({
            "Metric": ["Chargeable lines checked", "Matched", "Unmatched", "Ignored £0 lines", "Match percentage", "Run date/time"],
            "Value": [total, matched, unmatched, ignored, f"{match_pct}%", datetime.now().strftime("%d/%m/%Y %H:%M")],
        })
        excel_bytes = make_output_excel(checked_df, ignored_df, summary_df)
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        st.session_state["vendor_on_hire_results"] = {
            "checked_df": checked_df,
            "ignored_df": ignored_df,
            "summary_df": summary_df,
            "excel_bytes": excel_bytes,
            "total": total,
            "matched": matched,
            "unmatched": unmatched,
            "ignored": ignored,
            "match_pct": match_pct,
            "excel_filename": f"PAS_Vendor_On_Hire_Checked_{stamp}.xlsx",
        }
    except Exception as e:
        st.error(f"Something went wrong: {e}")
        st.exception(e)

results = st.session_state.get("vendor_on_hire_results")

if results is not None:
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M8 7V3h8l4 4v14H6V7z"/><path d="M16 3v5h5"/><path d="M9 13h6"/><path d="M9 17h4"/><path d="M4 7h2v14h12"/></svg></div><div><div class="kpi-label">Lines checked</div><div class="kpi-value">{results["total"]}</div><div class="kpi-sub">£0 lines ignored</div></div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="9"/><path d="M8 12.5l2.7 2.7L16.5 9"/></svg></div><div><div class="kpi-label">Matched</div><div class="kpi-value">{results["matched"]}</div><div class="kpi-sub">Still live/on hire</div></div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M12 3l10 18H2L12 3z"/><path d="M12 9v5"/><path d="M12 18h.01"/></svg></div><div><div class="kpi-label">Unmatched</div><div class="kpi-value">{results["unmatched"]}</div><div class="kpi-sub">Needs checking</div></div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="kpi-card"><div class="kpi-icon"><svg viewBox="0 0 24 24"><path d="M3 20h18"/><path d="M6 16v-4"/><path d="M11 16V8"/><path d="M16 16v-6"/><path d="M19 6l-5 5-3-3-5 5"/></svg></div><div><div class="kpi-label">Match %</div><div class="kpi-value">{results["match_pct"]}%</div><div class="kpi-sub">Core KPI</div></div></div>', unsafe_allow_html=True)

    st.markdown('<div class="pas-results-title">Results</div>', unsafe_allow_html=True)
    render_results_table(results["checked_df"])

    dl_left, dl_mid, dl_right = st.columns([1.3, 1, 1.3])
    with dl_mid:
        st.download_button(
            "⬇  Download Excel",
            data=results["excel_bytes"],
            file_name=results["excel_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
else:
    st.info("Upload vendor hire report and Material & Plant Orders, then click Run on-hire check.")


if "animation_shown" not in st.session_state:
    render_bottom_chase()
    st.session_state["animation_shown"] = True
